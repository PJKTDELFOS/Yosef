from openpyxl import load_workbook
from datetime import datetime
'''
contrato = models.ForeignKey(Contratos, on_delete=models.CASCADE, related_name='pedidos')
    numero=models.CharField(max_length=100)
    valor = models.DecimalField(max_digits=10, decimal_places=2)
    data_origem=models.DateField(blank=True,null=True)
    cnpj_contratante=models.CharField(max_length=18,default='')
    contratante=models.CharField(max_length=60,default='')
    empenho=models.CharField(max_length=50,default='')
    ordem_fornecimento=models.CharField(max_length=50,blank=True,default='')
    recebimento_empenho=models.DateField(blank=True,null=True)
    contato=models.CharField(max_length=50,blank=True,default='')
    telefone=models.CharField(max_length=50,blank=True,default='')
    email=models.EmailField(max_length=50,default='')
    objeto=models.TextField(blank=True,max_length=600,default='')
    data_entrega=models.DateTimeField(blank=True,null=True)
    unidade_fornecimento=models.CharField(max_length=50,default='')
    qtde=models.DecimalField(max_digits=10, decimal_places=2,default=0)
    coordenador=models.CharField(max_length=50,default='')
    show = models.BooleanField(default=True)
    status = models.CharField(default=None, max_length=25, choices=(
'''


def criar_pedido(instance):
    template_form = 'modelo_pedido.xlsx'
    workbook = load_workbook(filename=template_form)
    worksheet = workbook['sheet']
    worksheet['I9'] = instance.numero
    worksheet['I10'] = instance.data_origem
    worksheet['I12'] = instance.cnpj_contratante
    worksheet['B12'] = instance.contratante
    worksheet['A15'] = instance.contrato
    worksheet['C15'] = instance.empenho
    worksheet['D15'] = instance.ordem_fornecimento
    worksheet['E15'] = instance.data_origem
    worksheet['G15'] = instance.contato
    worksheet['H15'] = instance.telefone
    worksheet['I15'] = instance.email
    worksheet['D17'] = instance.objeto
    worksheet['B16'] = instance.data_entrega
    worksheet['D16'] = instance.local_entrega
    worksheet['A21'] = instance.unidade_fornecimento
    worksheet['B21'] = instance.qtde
    worksheet['A23'] = instance.observacoes
    worksheet['B51'] = instance.coordenador
    name=f'Pedido nº{instance.numero}-contrato:{instance.contrato} contratante:{instance.contrato}'
    workbook.save(filename=f'{name}.xlsx')
    print(f"Planilha salva como '{name}.xlsx'.")


def atualizar_pedido(name, newn_pedido, data_att, id, newcnpj=None, newcontratante=None,
                     newendereco_contratante=None, newn_contrato=None, newAta_RP=None,
                     newempenho=None, newo_f=None, newdata_recebimento_empenho=None, nf=None,
                     newcontato=None, newtelefone=None, newemail=None, newobjeto=None,
                     newdata_hora_entrega=None, newlocal_entrega=None, newU_F=None,
                     newqtde=None, newobs=None, newcoord=None,novo_status=None,novo_valor=None):
    filename = f'{name}.xlsx'
    workbook = load_workbook(filename=filename)
    worksheet = workbook['sheet']
    worksheet['I09'] = newn_pedido or worksheet['I09'].value
    worksheet['I11'] = data_att
    worksheet['I12'] = newcnpj or worksheet['I12'].value
    worksheet['B12'] = newcontratante or worksheet['B12'].value
    worksheet['B13'] = newendereco_contratante or worksheet['B13'].value
    worksheet['A15'] = newn_contrato or worksheet['A15'].value
    worksheet['B15'] = newAta_RP or worksheet['B15'].value
    worksheet['C15'] = newempenho or worksheet['C15'].value
    worksheet['D15'] = newo_f or worksheet['D15'].value
    worksheet['E15'] = newdata_recebimento_empenho or worksheet['E15'].value
    worksheet['F15'] = nf or worksheet['F15'].value
    worksheet['G15'] = newcontato or worksheet['G15'].value
    worksheet['H15'] = newtelefone or worksheet['H15'].value
    worksheet['I15'] = newemail or worksheet['I15'].value
    worksheet['D17'] = newobjeto or worksheet['D17'].value
    worksheet['B16'] = newdata_hora_entrega or worksheet['B16'].value
    worksheet['D16'] = newlocal_entrega or worksheet['D16'].value
    worksheet['A21'] = newU_F or worksheet['A21'].value
    worksheet['B21'] = newqtde or worksheet['B21'].value
    worksheet['A23'] = newobs or worksheet['A23'].value
    worksheet['B51'] = newcoord or worksheet['B51'].value
    workbook.save(filename=filename)
    print(f"Pedido '{name}.xlsx' atualizado com sucesso.")